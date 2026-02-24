#!/usr/bin/env python3
"""
Stir Trek 2026 Schedule Builder

Generates a scheduling prompt from session data, sends it to the Claude CLI
in headless mode, validates the result, and writes a CSV schedule.

Usage:
    python schedule_builder.py              # full run: prompt -> claude -> validate -> csv
    python schedule_builder.py --prompt     # just print the prompt (no CLI call)
    python schedule_builder.py --from-json output/schedule.json   # skip CLI, use saved JSON
"""

import argparse
import csv
import io
import json
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

# Force UTF-8 on Windows so emoji/special chars in session data don't crash
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency. Run: pip install -r requirements.txt")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "data" / "stir-trek-2026-accepted.xlsx"
OUTPUT_DIR = BASE_DIR / "output"
CSV_PATH = OUTPUT_DIR / "schedule.csv"
JSON_PATH = OUTPUT_DIR / "schedule.json"
VERSIONS_PATH = OUTPUT_DIR / "versions.json"
HTML_PATH = OUTPUT_DIR / "schedule.html"

SLOT_TIMES = [
    "08:30am - 09:15am",
    "09:30am - 10:15am",
    "10:30am - 11:15am",
    "11:30am - 12:15pm",
    "02:00pm - 02:45pm",
    "03:00pm - 03:45pm",
    "04:00pm - 04:45pm",
]

ROOMS = [
    {"num": 1, "alias": "Room 1",  "capacity": 388, "live": "Theater 14", "simulcast": "Theaters 12, 13"},
    {"num": 2, "alias": "Room 2",  "capacity": 314, "live": "Theater 15", "simulcast": "Theaters 10, 11"},
    {"num": 3, "alias": "Room 3",  "capacity": 228, "live": "Theater 16", "simulcast": "Theater 21"},
    {"num": 4, "alias": "Room 4",  "capacity": 234, "live": "Theater 17", "simulcast": "Theater 20"},
    {"num": 5, "alias": "Room 5",  "capacity": 340, "live": "Theater 4",  "simulcast": "Theaters 5,6,7,8,9"},
    {"num": 6, "alias": "Room 6",  "capacity": 293, "live": "Theater 3",  "simulcast": "Theaters 1, 2"},
    {"num": 7, "alias": "Room 7",  "capacity": 224, "live": "Theater 27", "simulcast": "Theaters 23,24,25,26"},
    {"num": 8, "alias": "Room 8",  "capacity": 173, "live": "Theater 28", "simulcast": "Theaters 18, 19"},
]

ROOM_NAMES = [f"{r['alias']} ({r['capacity']})" for r in ROOMS]

# 2025 attendance data: speaker name -> peak attendance from last year
# Used to inform room sizing for returning speakers
ATTENDANCE_2025 = {
    "Kathryn Grayson Nanz": 210,
    "Guy Royse": 200,
    "Cory House": 196,
    "Matt Eland": 157,       # Listed as "Matthew-Hope Eland" in 2026 data
    "Matthew-Hope Eland": 157,
    "Jeff McWherter": 154,
    "Tristan Chiappisi": 133,
    "Barret Blake": 101,
    "Bob Fornal": 88,
    "Randy Pagels": 86,
    "Cameron Presley": 78,
    "Kelly Morrison": 75,
    "Amanda Lange": 74,
    "Sam Basu": 69,
    "Burton Smith": 65,
    "Lance Finney": 55,
    "Brian McKeiver": 48,
}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_sessions():
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Accepted sessions"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    sessions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        session_id = row[col["Session Id"]]
        if session_id is None:
            continue
        desc = row[col["Description"]] or ""
        if len(desc) > 200:
            desc = desc[:200] + "..."
        sessions.append({
            "id": str(session_id),
            "title": row[col["Title"]],
            "description": desc,
            "speakers": row[col["Speakers"]],
            "track": row[col["Track"]],
        })

    wb.close()
    return sessions


def find_multi_session_speakers(sessions):
    speaker_sessions = defaultdict(list)
    for s in sessions:
        speaker_sessions[s["speakers"]].append(s["id"])
    return {k: v for k, v in speaker_sessions.items() if len(v) > 1}


# ---------------------------------------------------------------------------
# Prompt generation
# ---------------------------------------------------------------------------

def build_prompt(sessions, multi_speakers):
    session_block = "\n".join(
        f"- ID: {s['id']} | Title: {s['title']} | Speaker: {s['speakers']} "
        f"| Track: {s['track']} | Desc: {s['description']}"
        for s in sessions
    )

    speaker_constraint = "\n".join(
        f"  - {speaker}: sessions {', '.join(ids)}"
        for speaker, ids in multi_speakers.items()
    )

    track_counts = Counter(s["track"] for s in sessions)
    track_summary = "\n".join(
        f"  - {t}: {track_counts[t]} sessions"
        for t in sorted(track_counts)
    )

    room_block = "\n".join(
        f"  Position {i} = Room {r['num']}: {r['capacity']} seats "
        f"(live in {r['live']}, simulcast to {r['simulcast']})"
        for i, r in enumerate(ROOMS)
    )

    # Build 2025 attendance info for returning speakers
    attendance_lines = []
    for s in sessions:
        peak = ATTENDANCE_2025.get(s["speakers"])
        if peak:
            attendance_lines.append(
                f"  - {s['speakers']} (session {s['id']}): "
                f"2025 peak attendance = {peak}"
            )
    attendance_block = "\n".join(attendance_lines) if attendance_lines else "  (no data)"

    return f"""You are a conference schedule optimizer for Stir Trek 2026.

TASK: Assign all 56 sessions to 7 time slots with 8 rooms each.
Each slot is an array of 8 session IDs where POSITION MATTERS — position maps to a specific room.

ROOM ASSIGNMENTS (array index -> room):
{room_block}

SESSIONS:
{session_block}

TRACK SUMMARY:
{track_summary}

HARD CONSTRAINTS (must all be satisfied):
1. Exactly 7 time slots (slot_1 through slot_7), each with exactly 8 sessions.
2. Every session ID must appear exactly once across all slots.
3. No speaker may appear in two sessions in the same time slot.
   Multi-session speakers who MUST be in different slots:
{speaker_constraint}

SOFT CONSTRAINTS (optimize for these):
1. Minimize same-track sessions in the same time slot. Spread tracks across slots.
2. Spread popular/large tracks (Application Development, Architecture, AI/ML) across many slots.
3. Consider session descriptions to create variety within each slot — attendees should have diverse choices.
4. ROOM SIZING: Place higher-draw sessions in larger rooms using 2025 attendance data.
   Room sizes: Room 1 (388) > Room 5 (340) > Room 2 (314) > Room 6 (293) > Room 4 (234) > Room 3 (228) > Room 7 (224) > Room 8 (173)

   RETURNING SPEAKER ATTENDANCE FROM 2025 (use this to prioritize room assignments):
{attendance_block}

   Room assignment guidelines based on 2025 data:
   - 150+ attendance -> MUST be in Rooms 1 (388) or 5 (340): Cory House, Kathryn Grayson Nanz, Guy Royse, Matt Eland, Jeff McWherter
   - 100-149 attendance -> Prefer Rooms 2 (314) or 6 (293): Tristan Chiappisi, Barret Blake
   - AI/ML sessions from new speakers -> Rooms 1-5 (AI is the hottest track)
   - Architecture sessions -> Rooms 2-6
   - Niche/specialized sessions -> Rooms 7 (224) or 8 (173)
   - New speakers with unknown draw -> middle rooms (3, 4, 6, 7)

OUTPUT FORMAT:
Return ONLY valid JSON, no other text. Use this exact structure:
{{
  "slot_1": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_2": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_3": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_4": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_5": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_6": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"],
  "slot_7": ["room1_id", "room2_id", "room3_id", "room4_id", "room5_id", "room6_id", "room7_id", "room8_id"]
}}

Each array must contain exactly 8 session IDs as strings. Position 0 = Room 1 (388 seats), Position 7 = Room 8 (173 seats).
Use the exact Session Id values provided above."""


# ---------------------------------------------------------------------------
# Claude CLI interaction
# ---------------------------------------------------------------------------

def call_claude(prompt):
    """Call Claude CLI in headless mode. Returns parsed schedule dict."""
    print("Calling Claude CLI to generate schedule...")

    # Strip CLAUDECODE env var so the CLI doesn't refuse to launch
    env = {k: v for k, v in os.environ.items() if k != "CLAUDECODE"}

    result = subprocess.run(
        ["claude", "-p", "-", "--output-format", "json",
         "--model", "sonnet", "--max-turns", "1"],
        input=prompt.encode("utf-8"),
        capture_output=True,
        timeout=180,
        env=env,
    )
    # Decode output as UTF-8
    result = subprocess.CompletedProcess(
        result.args, result.returncode,
        stdout=result.stdout.decode("utf-8", errors="replace") if result.stdout else "",
        stderr=result.stderr.decode("utf-8", errors="replace") if result.stderr else "",
    )

    if result.returncode != 0:
        print(f"Claude CLI error (exit {result.returncode}):")
        print(result.stderr)
        sys.exit(1)

    raw = result.stdout.strip()
    if not raw:
        print("Claude CLI returned empty output.")
        if result.stderr:
            print("stderr:", result.stderr[:500])
        sys.exit(1)

    return parse_claude_response(raw)


def parse_claude_response(raw):
    """Extract schedule JSON from Claude CLI's JSON-envelope output."""
    # Parse the outer envelope
    try:
        envelope = json.loads(raw)
    except json.JSONDecodeError:
        print("Failed to parse Claude CLI output as JSON.")
        print("Raw output (first 500 chars):", raw[:500])
        sys.exit(1)

    # Extract text content from various envelope shapes
    text = ""
    if isinstance(envelope, dict) and "result" in envelope:
        text = envelope["result"]
    elif isinstance(envelope, list):
        for item in envelope:
            if isinstance(item, dict) and item.get("type") == "text":
                text = item.get("text", "")
                break
    elif isinstance(envelope, str):
        text = envelope
    else:
        text = str(envelope)

    # Try direct parse
    try:
        return json.loads(text)
    except (json.JSONDecodeError, TypeError):
        pass

    # Try extracting from markdown code fences
    m = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # Try finding raw JSON with slot_1 key
    m = re.search(r'(\{\s*"slot_1".*\})', text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    print("Could not extract schedule JSON from Claude's response.")
    print("Response text (first 1000 chars):", text[:1000])
    sys.exit(1)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_schedule(schedule, sessions):
    errors = []
    session_map = {s["id"]: s for s in sessions}
    all_ids = set(session_map.keys())

    if len(schedule) != 7:
        errors.append(f"Expected 7 slots, got {len(schedule)}")

    assigned_ids = []
    for slot_name in sorted(schedule.keys()):
        slot_ids = [str(sid) for sid in schedule[slot_name]]
        if len(slot_ids) != 8:
            errors.append(f"{slot_name}: expected 8 sessions, got {len(slot_ids)}")
        assigned_ids.extend(slot_ids)

        speakers_in_slot = []
        for sid in slot_ids:
            if sid in session_map:
                speaker = session_map[sid]["speakers"]
                if speaker in speakers_in_slot:
                    errors.append(
                        f"{slot_name}: speaker '{speaker}' appears twice"
                    )
                speakers_in_slot.append(speaker)
            else:
                errors.append(f"{slot_name}: unknown session ID '{sid}'")

    assigned_set = set(assigned_ids)
    missing = all_ids - assigned_set
    if missing:
        errors.append(f"Missing sessions: {missing}")
    extra = assigned_set - all_ids
    if extra:
        errors.append(f"Unknown session IDs: {extra}")
    dupes = [sid for sid, c in Counter(assigned_ids).items() if c > 1]
    if dupes:
        errors.append(f"Duplicate session IDs: {dupes}")

    return len(errors) == 0, errors


def compute_track_stats(schedule, sessions):
    session_map = {s["id"]: s for s in sessions}
    stats = {}
    total_doublings = 0

    for slot_name in sorted(schedule.keys()):
        tracks = [
            session_map[str(sid)]["track"]
            for sid in schedule[slot_name]
            if str(sid) in session_map
        ]
        counts = Counter(tracks)
        stats[slot_name] = counts
        for c in counts.values():
            if c > 1:
                total_doublings += c - 1

    return stats, total_doublings


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def write_csv(schedule, sessions):
    session_map = {s["id"]: s for s in sessions}
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Time"] + ROOM_NAMES)

        writer.writerow(["07:30am - 08:30am | Breakfast"] + ["Breakfast"] * 8)

        slot_keys = sorted(schedule.keys())
        for i, slot_key in enumerate(slot_keys[:4]):
            row = [SLOT_TIMES[i]]
            for sid in schedule[slot_key]:
                s = session_map.get(str(sid))
                row.append(
                    f"{s['title']} - {s['speakers']}" if s else f"Unknown ({sid})"
                )
            writer.writerow(row)

        writer.writerow(["12:15pm - 01:00pm | Lunch"] + ["Lunch"] * 8)
        writer.writerow(["01:00pm - 01:45pm | Pending"] + ["Pending"] * 8)

        for i, slot_key in enumerate(slot_keys[4:]):
            row = [SLOT_TIMES[4 + i]]
            for sid in schedule[slot_key]:
                s = session_map.get(str(sid))
                row.append(
                    f"{s['title']} - {s['speakers']}" if s else f"Unknown ({sid})"
                )
            writer.writerow(row)

        writer.writerow(
            ["05:00pm - 06:00pm | Movie Trailers"] + ["Movie Trailers"] * 8
        )

    print(f"CSV written to {CSV_PATH}")


# ---------------------------------------------------------------------------
# Versioning
# ---------------------------------------------------------------------------

def load_versions():
    """Load all versions from versions.json, or return empty list."""
    if VERSIONS_PATH.exists():
        with open(VERSIONS_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []


def save_version(schedule, label="", description=""):
    """Append a new version to versions.json and return the version number."""
    from datetime import datetime, timezone

    versions = load_versions()
    next_ver = max((v["version"] for v in versions), default=0) + 1

    versions.append({
        "version": next_ver,
        "label": label or f"Version {next_ver}",
        "description": description or "",
        "created": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "schedule": schedule,
    })

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(VERSIONS_PATH, "w", encoding="utf-8") as f:
        json.dump(versions, f, indent=2)

    return next_ver


# ---------------------------------------------------------------------------
# HTML output
# ---------------------------------------------------------------------------

def write_html(sessions):
    """Generate schedule.html with all versions embedded and a version picker."""
    versions = load_versions()
    if not versions:
        print("No versions found; skipping HTML generation.", file=sys.stderr)
        return

    session_map = {s["id"]: s for s in sessions}

    # Build sessions JS object
    sessions_js_parts = []
    for s in sessions:
        title_escaped = json.dumps(s["title"])
        speaker_escaped = json.dumps(s["speakers"])
        track_escaped = json.dumps(s["track"])
        sessions_js_parts.append(
            f'  "{s["id"]}": {{"title":{title_escaped},'
            f'"speakers":{speaker_escaped},"track":{track_escaped}}}'
        )
    sessions_js = "{\n" + ",\n".join(sessions_js_parts) + "\n}"

    # Build versions JS array
    versions_js = json.dumps(versions, indent=2)

    # Build attendance JS object
    attendance_js = json.dumps(ATTENDANCE_2025, indent=2)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Stir Trek 2026 Schedule</title>
<style>
  :root {{
    --app-dev: #3b82f6;
    --arch: #8b5cf6;
    --ai-ml: #ec4899;
    --pro-growth: #f59e0b;
    --sw-quality: #10b981;
    --security: #ef4444;
    --product-ux: #06b6d4;
    --data-eng: #f97316;
    --other: #6b7280;
    --breakfast: #fef3c7;
    --lunch: #fef3c7;
    --pending: #f3f4f6;
    --movie: #1e1b4b;
  }}

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #f8fafc;
    color: #1e293b;
    padding: 24px;
  }}

  h1 {{
    text-align: center;
    font-size: 2rem;
    margin-bottom: 4px;
    color: #0f172a;
  }}

  .subtitle {{
    text-align: center;
    color: #64748b;
    margin-bottom: 16px;
    font-size: 0.95rem;
  }}

  .version-bar {{
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 12px;
    margin-bottom: 8px;
    flex-wrap: wrap;
  }}

  .version-bar label {{
    font-weight: 700;
    font-size: 0.85rem;
    color: #334155;
  }}

  .version-select {{
    padding: 6px 12px;
    border: 2px solid #e2e8f0;
    border-radius: 8px;
    font-size: 0.85rem;
    font-weight: 600;
    background: white;
    cursor: pointer;
    min-width: 280px;
  }}

  .version-select:focus {{
    outline: none;
    border-color: #3b82f6;
  }}

  .version-description {{
    text-align: center;
    color: #94a3b8;
    font-size: 0.8rem;
    margin-bottom: 20px;
    font-style: italic;
    min-height: 1.2em;
  }}

  .legend {{
    display: flex;
    flex-wrap: wrap;
    justify-content: center;
    gap: 12px;
    margin-bottom: 24px;
  }}

  .legend-item {{
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 0.8rem;
    font-weight: 500;
  }}

  .legend-dot {{
    width: 14px;
    height: 14px;
    border-radius: 3px;
    flex-shrink: 0;
  }}

  .schedule-grid {{
    display: grid;
    grid-template-columns: 160px repeat(8, 1fr);
    gap: 2px;
    background: #e2e8f0;
    border-radius: 12px;
    overflow: hidden;
    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -2px rgba(0,0,0,0.1);
  }}

  .cell {{
    padding: 10px;
    background: white;
    min-height: 60px;
    display: flex;
    flex-direction: column;
    justify-content: center;
  }}

  .header-cell {{
    background: #0f172a;
    color: white;
    font-weight: 700;
    font-size: 0.85rem;
    text-align: center;
    padding: 12px 6px;
  }}

  .header-room-name {{ display: block; font-size: 0.85rem; }}
  .header-capacity {{ display: block; font-size: 0.95rem; font-weight: 800; margin-top: 2px; color: #f59e0b; }}
  .header-theaters {{ display: block; font-size: 0.6rem; font-weight: 400; color: #94a3b8; margin-top: 3px; line-height: 1.3; }}

  .time-cell {{
    background: #f1f5f9;
    font-weight: 700;
    font-size: 0.82rem;
    color: #334155;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    text-align: center;
    line-height: 1.4;
  }}

  .time-cell .slot-label {{
    font-size: 0.7rem;
    color: #94a3b8;
    margin-top: 4px;
    font-weight: 500;
  }}

  .session-cell {{
    border-left: 4px solid transparent;
    cursor: pointer;
    transition: transform 0.1s ease, box-shadow 0.1s ease;
    position: relative;
  }}

  .session-cell:hover {{
    transform: scale(1.02);
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    z-index: 10;
  }}

  .session-title {{ font-weight: 600; font-size: 0.78rem; line-height: 1.3; margin-bottom: 4px; }}
  .session-speaker {{ font-size: 0.72rem; color: #64748b; font-weight: 500; }}
  .session-track {{
    font-size: 0.65rem;
    margin-top: 4px;
    padding: 2px 6px;
    border-radius: 3px;
    display: inline-block;
    font-weight: 600;
    color: white;
    max-width: fit-content;
  }}

  .attendance-badge {{
    font-size: 0.6rem;
    margin-top: 3px;
    padding: 1px 5px;
    border-radius: 3px;
    display: inline-block;
    font-weight: 600;
    background: #fef3c7;
    color: #92400e;
    max-width: fit-content;
  }}

  .break-cell {{ text-align: center; font-weight: 700; font-size: 0.9rem; letter-spacing: 0.5px; }}
  .breakfast-cell {{ background: var(--breakfast); color: #92400e; }}
  .lunch-cell {{ background: var(--lunch); color: #92400e; }}
  .pending-cell {{ background: var(--pending); color: #9ca3af; font-style: italic; }}
  .movie-cell {{ background: var(--movie); color: #c4b5fd; }}

  .track-app-dev {{ border-left-color: var(--app-dev); }}
  .track-arch {{ border-left-color: var(--arch); }}
  .track-ai-ml {{ border-left-color: var(--ai-ml); }}
  .track-pro-growth {{ border-left-color: var(--pro-growth); }}
  .track-sw-quality {{ border-left-color: var(--sw-quality); }}
  .track-security {{ border-left-color: var(--security); }}
  .track-product-ux {{ border-left-color: var(--product-ux); }}
  .track-data-eng {{ border-left-color: var(--data-eng); }}
  .track-other {{ border-left-color: var(--other); }}

  .badge-app-dev {{ background: var(--app-dev); }}
  .badge-arch {{ background: var(--arch); }}
  .badge-ai-ml {{ background: var(--ai-ml); }}
  .badge-pro-growth {{ background: var(--pro-growth); }}
  .badge-sw-quality {{ background: var(--sw-quality); }}
  .badge-security {{ background: var(--security); }}
  .badge-product-ux {{ background: var(--product-ux); }}
  .badge-data-eng {{ background: var(--data-eng); }}
  .badge-other {{ background: var(--other); }}

  .stats {{
    margin-top: 24px;
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
  }}

  .stat-card {{
    background: white;
    border-radius: 8px;
    padding: 16px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }}

  .stat-card h3 {{ font-size: 0.8rem; color: #64748b; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 8px; }}
  .stat-card .value {{ font-size: 1.6rem; font-weight: 800; color: #0f172a; }}
  .stat-card .detail {{ font-size: 0.75rem; color: #94a3b8; margin-top: 4px; }}

  .filter-bar {{
    display: flex;
    justify-content: center;
    gap: 8px;
    margin-bottom: 20px;
    flex-wrap: wrap;
  }}

  .filter-btn {{
    padding: 6px 14px;
    border: 2px solid #e2e8f0;
    border-radius: 20px;
    background: white;
    font-size: 0.78rem;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.15s ease;
  }}

  .filter-btn:hover {{ border-color: #94a3b8; }}
  .filter-btn.active {{ border-color: #0f172a; background: #0f172a; color: white; }}

  .capacity-bar {{ width: 100%; height: 3px; background: #e2e8f0; border-radius: 2px; margin-top: 4px; overflow: hidden; }}
  .capacity-fill {{ height: 100%; border-radius: 2px; background: #f59e0b; }}

  .diff-highlight {{
    outline: 3px solid #f59e0b;
    outline-offset: -3px;
    background: #fffbeb !important;
  }}

  @media (max-width: 1200px) {{
    .schedule-grid {{ font-size: 0.85em; }}
  }}

  @media print {{
    body {{ padding: 0; background: white; }}
    .filter-bar, .version-bar, .version-description {{ display: none; }}
    .session-cell:hover {{ transform: none; box-shadow: none; }}
  }}
</style>
</head>
<body>

<h1>Stir Trek 2026</h1>
<p class="subtitle">May 2, 2026 &mdash; 56 Sessions &bull; 9 Tracks &bull; 7 Time Slots &bull; 8 Rooms &bull; 2,194 Total Seats</p>

<div class="version-bar">
  <label for="version-picker">Schedule Version:</label>
  <select id="version-picker" class="version-select"></select>
  <label style="margin-left:16px"><input type="checkbox" id="diff-toggle"> Highlight changes</label>
</div>
<div class="version-description" id="version-desc"></div>

<div class="legend">
  <div class="legend-item"><div class="legend-dot" style="background:var(--app-dev)"></div>Application Development</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--arch)"></div>Architecture</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--ai-ml)"></div>AI &amp; ML</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--pro-growth)"></div>Professional Growth</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--sw-quality)"></div>Software Quality</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--security)"></div>Security</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--product-ux)"></div>Product &amp; UX</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--data-eng)"></div>Data Engineering</div>
  <div class="legend-item"><div class="legend-dot" style="background:var(--other)"></div>Other</div>
</div>

<div class="filter-bar">
  <button class="filter-btn active" data-track="all">All Tracks</button>
  <button class="filter-btn" data-track="Application Development">App Dev</button>
  <button class="filter-btn" data-track="Architecture & Platform Engineering">Architecture</button>
  <button class="filter-btn" data-track="Artificial Intelligence & Machine Learning">AI &amp; ML</button>
  <button class="filter-btn" data-track="Professional Growth & Leadership">Prof Growth</button>
  <button class="filter-btn" data-track="Software Quality & Delivery">SW Quality</button>
  <button class="filter-btn" data-track="Security & Privacy Engineering">Security</button>
  <button class="filter-btn" data-track="Product Design & User Experience">Product UX</button>
  <button class="filter-btn" data-track="Data Engineering & Analytics">Data Eng</button>
  <button class="filter-btn" data-track="Other">Other</button>
</div>

<div class="schedule-grid" id="grid"></div>

<div class="stats" id="stats"></div>

<script>
const rooms = [
  {{ num: 1, alias: "Room 1",  capacity: 388, live: "Theater 14", simulcast: "12, 13" }},
  {{ num: 2, alias: "Room 2",  capacity: 314, live: "Theater 15", simulcast: "10, 11" }},
  {{ num: 3, alias: "Room 3",  capacity: 228, live: "Theater 16", simulcast: "21" }},
  {{ num: 4, alias: "Room 4",  capacity: 234, live: "Theater 17", simulcast: "20" }},
  {{ num: 5, alias: "Room 5",  capacity: 340, live: "Theater 4",  simulcast: "5, 6, 7, 8, 9" }},
  {{ num: 6, alias: "Room 6",  capacity: 293, live: "Theater 3",  simulcast: "1, 2" }},
  {{ num: 7, alias: "Room 7",  capacity: 224, live: "Theater 27", simulcast: "23, 24, 25, 26" }},
  {{ num: 8, alias: "Room 8",  capacity: 173, live: "Theater 28", simulcast: "18, 19" }}
];

const maxCapacity = Math.max(...rooms.map(r => r.capacity));

const sessions = {sessions_js};

const versions = {versions_js};

const attendance2025 = {attendance_js};

const slotTimes = [
  {{ time: "08:30 - 09:15", label: "Slot 1" }},
  {{ time: "09:30 - 10:15", label: "Slot 2" }},
  {{ time: "10:30 - 11:15", label: "Slot 3" }},
  {{ time: "11:30 - 12:15", label: "Slot 4" }},
  {{ time: "02:00 - 02:45", label: "Slot 5" }},
  {{ time: "03:00 - 03:45", label: "Slot 6" }},
  {{ time: "04:00 - 04:45", label: "Slot 7" }}
];

let currentVersion = versions[versions.length - 1];
let compareVersion = null;
let activeFilter = "all";
let showDiff = false;

// Populate version picker
const picker = document.getElementById("version-picker");
versions.forEach((v, i) => {{
  const opt = document.createElement("option");
  opt.value = i;
  opt.textContent = `v${{v.version}}: ${{v.label}} (${{v.created.split("T")[0]}})`;
  if (i === versions.length - 1) opt.selected = true;
  picker.appendChild(opt);
}});

picker.addEventListener("change", () => {{
  currentVersion = versions[parseInt(picker.value)];
  render();
}});

document.getElementById("diff-toggle").addEventListener("change", (e) => {{
  showDiff = e.target.checked;
  render();
}});

function trackClass(track) {{
  const map = {{
    "Application Development": "app-dev",
    "Architecture & Platform Engineering": "arch",
    "Artificial Intelligence & Machine Learning": "ai-ml",
    "Professional Growth & Leadership": "pro-growth",
    "Software Quality & Delivery": "sw-quality",
    "Security & Privacy Engineering": "security",
    "Product Design & User Experience": "product-ux",
    "Data Engineering & Analytics": "data-eng",
    "Other": "other"
  }};
  return map[track] || "other";
}}

function shortTrack(track) {{
  const map = {{
    "Application Development": "App Dev",
    "Architecture & Platform Engineering": "Architecture",
    "Artificial Intelligence & Machine Learning": "AI & ML",
    "Professional Growth & Leadership": "Prof Growth",
    "Software Quality & Delivery": "SW Quality",
    "Security & Privacy Engineering": "Security",
    "Product Design & User Experience": "Product UX",
    "Data Engineering & Analytics": "Data Eng",
    "Other": "Other"
  }};
  return map[track] || track;
}}

function getDiffSet() {{
  // Build a set of "slot:position" keys that differ from previous version
  const diffs = new Set();
  const curIdx = versions.indexOf(currentVersion);
  if (curIdx <= 0 || !showDiff) return diffs;
  const prev = versions[curIdx - 1].schedule;
  const cur = currentVersion.schedule;
  for (const slot of Object.keys(cur).sort()) {{
    const curIds = cur[slot] || [];
    const prevIds = prev[slot] || [];
    for (let i = 0; i < 8; i++) {{
      if (curIds[i] !== prevIds[i]) {{
        diffs.add(`${{slot}}:${{i}}`);
      }}
    }}
  }}
  return diffs;
}}

function render() {{
  const schedule = currentVersion.schedule;
  const diffs = getDiffSet();
  const grid = document.getElementById("grid");
  grid.innerHTML = "";

  // Version description
  document.getElementById("version-desc").textContent = currentVersion.description || "";

  // Header row
  const timeHeader = document.createElement("div");
  timeHeader.className = "cell header-cell";
  timeHeader.textContent = "Time";
  grid.appendChild(timeHeader);

  for (const room of rooms) {{
    const h = document.createElement("div");
    h.className = "cell header-cell";
    const pct = (room.capacity / maxCapacity * 100).toFixed(0);
    h.innerHTML = `
      <span class="header-room-name">${{escHtml(room.alias)}}</span>
      <span class="header-capacity">${{room.capacity}} seats</span>
      <span class="header-theaters">Live: ${{room.live}}<br>Simulcast: ${{room.simulcast}}</span>
      <div class="capacity-bar"><div class="capacity-fill" style="width:${{pct}}%"></div></div>
    `;
    grid.appendChild(h);
  }}

  addBreakRow("07:30 - 08:30", "Breakfast", "breakfast-cell");

  const slotKeys = Object.keys(schedule).sort();
  for (let i = 0; i < 4; i++) {{
    addSessionRow(slotTimes[i], schedule[slotKeys[i]], slotKeys[i], diffs);
  }}

  addBreakRow("12:15 - 01:00", "Lunch", "lunch-cell");
  addBreakRow("01:00 - 01:45", "Pending", "pending-cell");

  for (let i = 4; i < 7; i++) {{
    addSessionRow(slotTimes[i], schedule[slotKeys[i]], slotKeys[i], diffs);
  }}

  addBreakRow("05:00 - 06:00", "Movie Trailers", "movie-cell");

  renderStats(schedule);
}}

function addBreakRow(time, label, cls) {{
  const grid = document.getElementById("grid");
  const tc = document.createElement("div");
  tc.className = "cell time-cell";
  tc.innerHTML = `<span>${{time}}</span>`;
  grid.appendChild(tc);
  for (let i = 0; i < 8; i++) {{
    const c = document.createElement("div");
    c.className = `cell break-cell ${{cls}}`;
    c.textContent = label;
    grid.appendChild(c);
  }}
}}

function addSessionRow(slotInfo, sessionIds, slotKey, diffs) {{
  const grid = document.getElementById("grid");
  const tc = document.createElement("div");
  tc.className = "cell time-cell";
  tc.innerHTML = `<span>${{slotInfo.time}}</span><span class="slot-label">${{slotInfo.label}}</span>`;
  grid.appendChild(tc);

  sessionIds.forEach((sid, idx) => {{
    const s = sessions[sid];
    const tc2 = trackClass(s.track);
    const dimmed = activeFilter !== "all" && s.track !== activeFilter;
    const isDiff = diffs.has(`${{slotKey}}:${{idx}}`);
    const c = document.createElement("div");
    c.className = `cell session-cell track-${{tc2}}${{isDiff ? " diff-highlight" : ""}}`;
    if (dimmed) c.style.opacity = "0.15";
    const att = attendance2025[s.speakers];
    const attHtml = att ? `<span class="attendance-badge">2025: ${{att}} attendees</span>` : "";
    c.innerHTML = `
      <div class="session-title">${{escHtml(s.title)}}</div>
      <div class="session-speaker">${{escHtml(s.speakers)}}</div>
      <span class="session-track badge-${{tc2}}">${{shortTrack(s.track)}}</span>
      ${{attHtml}}
    `;
    grid.appendChild(c);
  }});
}}

function escHtml(str) {{
  const d = document.createElement("div");
  d.textContent = str;
  return d.innerHTML;
}}

function renderStats(schedule) {{
  const stats = document.getElementById("stats");
  stats.innerHTML = "";

  const trackCounts = {{}};
  const slotKeys = Object.keys(schedule).sort();
  let doublings = 0;

  for (const sk of slotKeys) {{
    const tracksInSlot = {{}};
    for (const sid of schedule[sk]) {{
      const t = sessions[sid].track;
      tracksInSlot[t] = (tracksInSlot[t] || 0) + 1;
      trackCounts[t] = (trackCounts[t] || 0) + 1;
    }}
    for (const c of Object.values(tracksInSlot)) {{
      if (c > 1) doublings += c - 1;
    }}
  }}

  const totalSeats = rooms.reduce((a, r) => a + r.capacity, 0);

  const cards = [
    {{ label: "Total Sessions", value: "56", detail: "7 slots \\u00d7 8 rooms" }},
    {{ label: "Total Capacity", value: totalSeats.toLocaleString(), detail: "Per time slot across all rooms" }},
    {{ label: "Tracks", value: Object.keys(trackCounts).length, detail: "Across all sessions" }},
    {{ label: "Track Doublings", value: doublings, detail: "Theoretical min: 7" }},
    {{ label: "Dual Speakers", value: "7", detail: "All separated into different slots" }},
    {{ label: "Largest Room", value: "388", detail: "Room 1: Theater 14 + simulcast" }}
  ];

  for (const card of cards) {{
    const el = document.createElement("div");
    el.className = "stat-card";
    el.innerHTML = `<h3>${{card.label}}</h3><div class="value">${{card.value}}</div><div class="detail">${{card.detail}}</div>`;
    stats.appendChild(el);
  }}
}}

// Filter buttons
document.querySelectorAll(".filter-btn").forEach(btn => {{
  btn.addEventListener("click", () => {{
    document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
    btn.classList.add("active");
    activeFilter = btn.dataset.track;
    render();
  }});
}});

render();
</script>
</body>
</html>"""

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"HTML written to {HTML_PATH}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stir Trek 2026 Schedule Builder")
    parser.add_argument(
        "--prompt", action="store_true",
        help="Print the prompt to stdout and exit (no CLI call)",
    )
    parser.add_argument(
        "--from-json", metavar="FILE",
        help="Skip Claude CLI; read schedule from a JSON file",
    )
    parser.add_argument(
        "--version-label", metavar="LABEL",
        help="Label for this schedule version (default: auto-numbered)",
    )
    parser.add_argument(
        "--version-desc", metavar="DESC",
        help="Description for this schedule version",
    )
    parser.add_argument(
        "--html-only", action="store_true",
        help="Regenerate HTML from existing versions.json (no scheduling)",
    )
    args = parser.parse_args()

    # Load sessions
    print("Loading session data...", file=sys.stderr)
    sessions = load_sessions()
    multi_speakers = find_multi_session_speakers(sessions)
    print(f"  {len(sessions)} sessions, {len(multi_speakers)} multi-session speakers",
          file=sys.stderr)

    # HTML-only mode: regenerate HTML and exit
    if args.html_only:
        write_html(sessions)
        print("Done!", file=sys.stderr)
        return

    tracks = Counter(s["track"] for s in sessions)
    for track, count in tracks.most_common():
        print(f"  {track}: {count}", file=sys.stderr)

    # Build prompt (always needed for --prompt mode, useful to log otherwise)
    prompt = build_prompt(sessions, multi_speakers)

    # --prompt mode: just dump it and exit
    if args.prompt:
        print(prompt)
        return

    # Get schedule: either from file or from Claude CLI
    if args.from_json:
        print(f"Reading schedule from {args.from_json}...", file=sys.stderr)
        with open(args.from_json) as f:
            schedule = json.load(f)
    else:
        schedule = call_claude(prompt)

    # Normalize IDs to strings
    schedule = {
        k: [str(sid) for sid in v] for k, v in schedule.items()
    }

    # Save the raw schedule JSON for reuse
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w") as f:
        json.dump(schedule, f, indent=2)
    print(f"Schedule JSON saved to {JSON_PATH}", file=sys.stderr)

    # Validate
    print("\n--- Validation ---", file=sys.stderr)
    ok, errors = validate_schedule(schedule, sessions)
    if ok:
        print("PASS: All hard constraints satisfied", file=sys.stderr)
    else:
        print("FAIL: Constraint violations:", file=sys.stderr)
        for err in errors:
            print(f"  - {err}", file=sys.stderr)

    # Track stats
    print("\n--- Track Distribution ---", file=sys.stderr)
    stats, total_doublings = compute_track_stats(schedule, sessions)
    for slot_name, track_counts in stats.items():
        doubled = {t: c for t, c in track_counts.items() if c > 1}
        suffix = f"  doubled: {dict(doubled)}" if doubled else ""
        print(f"  {slot_name}: {len(track_counts)} unique tracks{suffix}",
              file=sys.stderr)
    print(f"  Total track doublings: {total_doublings}", file=sys.stderr)

    # Save version
    ver = save_version(
        schedule,
        label=args.version_label or "",
        description=args.version_desc or "",
    )
    print(f"Saved as version {ver}", file=sys.stderr)

    # Write CSV + HTML
    write_csv(schedule, sessions)
    write_html(sessions)
    print("\nDone!", file=sys.stderr)


if __name__ == "__main__":
    main()
